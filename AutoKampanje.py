import openpyxl
import tkinter as tk
import pygetwindow as gw
import pyautogui
import time
import json
import csv
from tkinter import filedialog, ttk
from PIL import Image, ImageTk
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Read the configuration file
with open('./data/config.json', 'r') as config_file:
    config = json.load(config_file)

# Access individual configuration values
target_program_title = config["target_program_title"]
power_butikk = config["power_butikk"]
logo_path = config["logo_path"]
file_path = config["file_path"]
ico_path = config["ico_path"]
search_bar_path = config["search_bar_path"]
info_bar_path = config["info_bar_path"]
maler_path = config["maler_path"]
delay_start = config["delay_start"]
pause_autogui = config["pause_autogui"]
x_width = config["x_width"]
#y_height = config["y_height"]
y_height = 520
search_plus_x = config["search_plus_x"]
search_plus_y = config["search_plus_y"]
info_plus_x = config["info_plus_x"]
info_plus_y = config["info_plus_y"]
ean_array = []
forpris_array = []
kampris_array= []
ean_csv = []

ikke_array_A = []
ikke_array_B = []
ikke_array_C = []

csv_count = 0
excel_count = 0
row = 2  
column_ean = 3 
column_forpris = 7
column_kampris = 6
bilde_status = "nei"


def browse_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
    if file_path:
        excel_file_label.config(text=file_path)
        excel_data()

def browse_csv_file():
    file_path = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if file_path:
        csv_file_label.config(text=file_path)
        csv_data()


def excel_data():
    global ean_array, forpris_array, kampris_array, row, excel_count
    excel_file_path = excel_file_label.cget("text") 
    excel_count = 0

    if excel_file_path:
        wb = openpyxl.load_workbook(excel_file_path)
        ws = wb.active
        while True:
            cell_value_A = ws.cell(row=row, column=column_ean).value
            cell_value_B = ws.cell(row=row, column=column_forpris).value
            cell_value_C = ws.cell(row=row, column=column_kampris).value
            if cell_value_A is not None and cell_value_A == str(cell_value_A):
                ean_array.append(str(cell_value_A))
            if cell_value_B is not None and cell_value_B == int(cell_value_B):
                forpris_array.append(int(cell_value_B))
            if cell_value_C is not None and cell_value_C == int(cell_value_C):
                kampris_array.append(int(cell_value_C))
            if cell_value_A is None or cell_value_B is None or cell_value_C is None:
                break
            row += 1

    excel_count = len(ean_array)
    excel_label.config(text=f"{excel_count} på kampanje", fg="#f15c25")


def csv_data():
    global ean_csv, csv_count
    csv_file_path = csv_file_label.cget("text")
    csv_count = 0
    ean_csv = []
    if csv_file_path:
        with open(csv_file_path, 'r') as csv_file:
            csv_reader = csv.DictReader(csv_file, delimiter=';') 

            for row in csv_reader:
                varenummer = row['Varenummer']
                ean_csv.append(varenummer)
    csv_count = len(ean_csv)
    csv_label.config(text=f"{csv_count} på lager", fg="#F1C225")
        

def csv_merge():
    global ean_array, forpris_array, kampris_array, merged_count
    global ikke_array_A,ikke_array_B,ikke_array_C, ikke_count
    merged_count = 0
    ikke_count = 0
    updated_values_array_A = []
    updated_values_array_B = []
    updated_values_array_C = []
    ikke_array_A = []
    ikke_array_B = []
    ikke_array_C = []

    for index, item in enumerate(ean_array):
        if item in ean_csv:
            updated_values_array_A.append(item)
            updated_values_array_B.append(forpris_array[index])
            updated_values_array_C.append(kampris_array[index])
        elif item not in ean_csv:
            ikke_array_A.append(item)
            ikke_array_B.append(forpris_array[index])
            ikke_array_C.append(kampris_array[index])
    ean_array = updated_values_array_A
    forpris_array = updated_values_array_B
    kampris_array = updated_values_array_C
    merged_count = len(ean_array)
    ikke_count = len(ikke_array_A)


def pris_sjekk():
    global ean_array, forpris_array, kampris_array
    temp_a=[]
    temp_b=[]
    temp_c=[]
    for value_A, value_B, value_C in zip(ean_array,forpris_array, kampris_array):
        # Check if the values are not equal
        if value_B != value_C:
            # If they are not equal, add them to the filtered arrays
            temp_a.append(value_A)
            temp_b.append(value_B)
            temp_c.append(value_C)

            ean_array = temp_a
            forpris_array = temp_b
            kampris_array = temp_c

def ready():
    global excel_count,csv_count,merged_count
    if csv_count != 0:
        csv_merge()
    enable_button2()
    pris_sjekk()

    ready_count = len(ean_array)
    success_label.config(text=f"{ready_count} varer på kampanje & beholdning", fg="#24db4f")


def reset():
    global excel_file_path, excel_count, ean_array, forpris_array, kampris_array
    global csv_file_path, csv_count, ean_csv, row
    excel_file_path = ""
    csv_file_path = ""
    ean_array = []
    forpris_array = []
    kampris_array = []
    ean_csv = []
    excel_count = 0
    csv_count = 0
    row = 2
    excel_file_label.config(text="")
    csv_file_label.config(text="")
    excel_label.config(text="")
    csv_label.config(text="")
    success_label.config(text="")

def save_data_excel():
    global file_name
    if len(ean_array) < 1:
        return status_label.config(text="No Data to save!", fg="#d51f2f")
    wb = openpyxl.Workbook()
    default_file_name = "kampanje_lager.xlsx"

  # Create the first sheet
    sheet1 = wb.active
    sheet1.title = "Kamp varer på lager"
    headers= ['Varekode', 'Kampanje Pris', 'Før Pris']
    sheet1.append(headers)
    for a1, a2, a3 in zip(ean_array, kampris_array, forpris_array):
        sheet1.append([a1, a2, a3])

    # Create the second sheet
    sheet2 = wb.create_sheet(title="Kamp varer ikke på lager")
    sheet2.append(headers)
    for b1, b2, b3 in zip(ikke_array_A, ikke_array_B, ikke_array_C):
        sheet2.append([b1, b2, b3])

    file_name = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],initialfile=default_file_name)
    if file_name:
        wb.save(file_name)
        style_excel_file(file_name)
        success_label.config(text="Data saved successfully")
    else:
        success_label.config(text="Data was not saved.")


def style_excel_file(excel_filename):
    column_widths = [30, 20, 20]
    column_colors = ["95B3D7", "f15c25", "B1A0C7"]
    font_sizes = [11, 11, 11]

    # Load the Excel file
    book = load_workbook(excel_filename)

    # Loop through all sheets in the workbook
    for sheet_name in book.sheetnames:
        sheet = book[sheet_name]

        # Adjust column width
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width

        # Apply different background color, font size, and left alignment to each column
        for col_idx, (color, font_size) in enumerate(zip(column_colors, font_sizes), start=1):
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    cell.font = Font(size=font_size)
                    cell.alignment = Alignment(horizontal='left', vertical='center')

        # Apply red background to rows with matching "kampanje pris" and "før pris" values
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            kampanje_value = row[1].value
            før_value = row[2].value
            if kampanje_value == før_value:
                for cell in row:
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Apply all borders to the data
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            for cell in row:
                cell.border = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin")
                )
    # Save the modified workbook
    book.save(excel_filename)

def on_combobox_select(event):
    global selected_value
    selected_value = select_box.get()
    print(selected_value)

def update_checkbox_label():
    if checkbox_var.get():
        bilde.set("ja")
        checkbox_label.config(text="Med bilde", fg="white")
    else:
        bilde.set("nei")
        checkbox_label.config(text="Uten bilde", fg="white")

active_window = gw.getActiveWindow()
screen_width, screen_height = pyautogui.size()
try:
    maler_button = pyautogui.locateOnScreen(maler_path)
except:
    print("maler png")

def main():
    window_title = target_program_title 
    window = gw.getWindowsWithTitle(window_title)
    active_window = gw.getActiveWindow()

    if (len(ean_array)) < 1:
        status_label.config(text="There is nothing to print!", fg="#d51f2f")
            
    if window:
            window = window[0]
            window.activate()  
            window.maximize()  

    active_window = gw.getActiveWindow()

    if active_window is None or active_window.title != target_program_title:
        status_label.config(text="Start Shoppa !", fg="#d51f2f", font=(9))
        root.after(4000, hide_status_message)

    else:
        time.sleep(delay_start)
        if maler_button is not None:
            pyautogui.click(maler_button)
        else:
            pyautogui.click(26,57)

        pyautogui.click(80,95)
        if bilde_status == "ja":
            pyautogui.typewrite("pris og bilde")
        elif bilde_status == "nei":
            pyautogui.typewrite("uten bilde")
        time.sleep(delay_start)
        pyautogui.click(90,150)
        time.sleep(delay_start)
        pyautogui.moveTo(100,777)
        if selected_value==options[0]:
            pyautogui.scroll(-5000)
            pyautogui.click(110,933)
        elif selected_value == options[1]:
            pyautogui.scroll(-5000)    
            pyautogui.click(111,861)
        elif selected_value == options[2]:
            pyautogui.scroll(-5000)
            pyautogui.click(106,791)
        elif selected_value == options[3]:
            pyautogui.scroll(5000)
            pyautogui.click(107,805)
        elif selected_value == options[4]:
            pyautogui.scroll(5000)
            pyautogui.click(104,601)
        elif selected_value == options[5]:
            pyautogui.scroll(5000)
            pyautogui.click(99,812)   
        time.sleep(delay_start)
        for index in range(len(ean_array)):
            item_A = ean_array[index]
            item_C = forpris_array[index]   
            item_B = kampris_array[index]
            
            search_bar = pyautogui.locateOnScreen(search_bar_path)
            if search_bar is not None:
                center_x = search_bar.left + (search_bar.width / 2)
                center_y = search_bar.top + (search_bar.height / 2)
                pyautogui.moveTo(center_x + search_plus_x, center_y - search_plus_y)
            else:
                pyautogui.moveTo(x=1043, y=56)
            time.sleep(delay_start)  
            pyautogui.doubleClick()
            pyautogui.typewrite(str(item_A))
            pyautogui.hotkey("enter")
            time.sleep(delay_start)
    
            x_pixel = int(screen_width * 0.900)
            y_pixel = int(screen_height * 0.275)
            pixel_color = pyautogui.pixel(x_pixel, y_pixel)
            time.sleep(delay_start)

            reference_color = (255, 255, 255)  
            if pixel_color == reference_color:
                pyautogui.moveTo(screen_width * 0.945, screen_height * 0.275)
            else:
                pyautogui.moveTo(screen_width * 0.945, screen_height * 0.175)

            pyautogui.mouseDown(button='left')
            pyautogui.mouseDown(button='right')
            pyautogui.moveTo(screen_width / 2, screen_height / 2) 
            pyautogui.mouseUp(button='left')
            pyautogui.mouseUp(button='right')
            
            if selected_value == options[0] or selected_value == options[1] or selected_value == options[2]:
                pyautogui.moveTo(1260,582)
            elif selected_value == options[1]:
                pyautogui.moveTo(1071,770) 
        
            pyautogui.click()
            pyautogui.typewrite(str(item_B))
            if item_B != item_C:
                if selected_value == options[1]:
                    pyautogui.hotkey('tab')
                    pyautogui.hotkey('tab')
                else:
                    pyautogui.hotkey('tab')
                pyautogui.typewrite(str(item_C))
            time.sleep(delay_start)
            info_bar = pyautogui.locateOnScreen(info_bar_path)
            search_x = info_bar.left + (info_bar.width / 2)
            search_y = info_bar.top + (info_bar.height / 2)

            pyautogui.moveTo(search_x - info_plus_x, search_y + info_plus_y)
            pyautogui.doubleClick()
            pyautogui.typewrite("5")
            time.sleep(delay_start)
                    
            pyautogui.hotkey('ctrl', 'q')
            
        pyautogui.hotkey('ctrl', 'p')

        
        root.after(1, hide_success_message)
        status_label.config(text=f"Program Execution Complete", fg="#24db4f")
        root.after(5000, hide_status_message)

        window2 = gw.getWindowsWithTitle(power_butikk)
        window2 = window2[0]  
        window2.activate()  


def hide_status_message():
    status_label.config(text="")   

def hide_success_message():
    success_label.config(text="")  

def enable_button2():
    run_button.config(state=tk.NORMAL)

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    window.geometry(f"{width}x{height}+{x}+{y}")

def on_enter(event):
    widget = event.widget  
    if widget['state'] == tk.NORMAL:
        widget.configure(
            background="white"#F1C225
        )

def on_leave(event):
    widget = event.widget  
    widget.configure(
        background="#f15c25"
    )

def on_enter2(event):
    widget = event.widget  
    if widget['state'] == tk.NORMAL:
        widget.configure(
            background="white"#f15c25
        )

def on_leave2(event):
    widget = event.widget  
    widget.configure(
        background="#F1C225"
    )

def on_enter4(event):
    widget = event.widget  
    if widget['state'] == tk.NORMAL:
        widget.configure(
            fg="#F1C225"
        )

def on_leave4(event):
    widget = event.widget  
    widget.configure(
        fg="white"
    )

root = tk.Tk()
root.title(power_butikk)
center_window(root, x_width, y_height)
root.iconbitmap(ico_path)
root.configure(bg="#2b2e33")
root.attributes("-alpha", 0.98)
root.resizable(False,False)


label_style = {
    "bg":"#2b2e33", 
    "font":("Roboto", 10),
}

logo_image = Image.open(logo_path)
logo_image = logo_image.resize((230, 75))
logo_photo = ImageTk.PhotoImage(logo_image)
logo_label = tk.Label(root, image=logo_photo, bg="#2b2e33")
logo_label.image = logo_photo  
logo_label.pack(pady = (10,0)) 


auto_shoppa = tk.Label(root)
auto_shoppa.configure(text="Kampanje", fg="white", bg="#2b2e33", font=("Roboto", 10, "bold italic"))
auto_shoppa.pack(pady=(0,35))
auto_shoppa.bind("<Enter>", on_enter4)
auto_shoppa.bind("<Leave>", on_leave4)


button_frame = tk.Frame(root)
button_frame.configure(background="#2b2e33")
button_frame.pack()

browse_button = tk.Button(
    button_frame,
    text="Excel",
    bg="#f15c25",
    fg="black",
    font=("Roboto",10,"bold"),
    cursor="hand2",
    command=browse_excel_file,
    relief=tk.FLAT,
    borderwidth=1,
    width=9,
    height=2,
)
browse_button.configure(cursor="hand2")
browse_button.pack(side=tk.LEFT, padx=1)
browse_button.bind("<Enter>", on_enter)
browse_button.bind("<Leave>", on_leave)

browse_csv_button = tk.Button(
    button_frame,
    text="Csv",
    bg="#F1C225",
    fg="black",
    font=("Roboto",10,"bold"),
    cursor="hand2",
    command=browse_csv_file,
    relief=tk.FLAT,
    borderwidth=1,
    width=7,
    height=2,
)

browse_csv_button.configure(cursor="hand2")
browse_csv_button.pack(side=tk.LEFT, padx=0)
browse_csv_button.bind("<Enter>", on_enter2)
browse_csv_button.bind("<Leave>", on_leave2)

label_buttom_frame = tk.Frame(root, bg="#2b2e33")
label_buttom_frame.pack(pady=(2, 5))

excel_label = tk.Label(label_buttom_frame, text="", fg="green", **label_style)
excel_label.grid(row=1, column=0, padx=5, pady=(0))
excel_file_label = tk.Label(root, text="")

csv_label = tk.Label(label_buttom_frame, text="", fg="green", **label_style)
csv_label.grid(row=1, column=1, padx=5, pady=(0))
csv_file_label = tk.Label(root, text="")

button2_frame = tk.Frame(root)
button2_frame.configure(background="#2b2e33")
button2_frame.pack(pady=5)

ready_button = tk.Button(
    button2_frame,
    text="Ready",
    bg="#f15c25",
    fg="black",
    font=("Roboto",10,"bold"),
    cursor="hand2",
    command=ready,
    relief=tk.FLAT,  
    borderwidth=1,
    width=9,
    height=2,  
)
ready_button.configure(cursor="hand2")
ready_button.pack(pady = 5,side=tk.LEFT, padx=1)
ready_button.bind("<Enter>", on_enter)
ready_button.bind("<Leave>", on_leave)

reset_button = tk.Button(
    button2_frame,
    text="Reset",
    bg="#F1C225",
    fg="black",
    font=("Roboto",10,"bold"),
    cursor="hand2",
    command=reset,
    relief=tk.FLAT,  
    borderwidth=1,
    width=7,
    height=2,  
)
reset_button.configure(cursor="hand2")
reset_button.pack(pady = 5,side=tk.LEFT, padx=0)
reset_button.bind("<Enter>", on_enter2)
reset_button.bind("<Leave>", on_leave2)

success_label = tk.Label(root, text="", fg="green", **label_style)
success_label.configure(cursor="hand2",relief=tk.FLAT,)
success_label.pack(pady = (0,5))


frame_mal = tk.Frame(root,bg="#373d41")
frame_mal.pack(pady=5, padx=1)

checkbox_var = tk.BooleanVar()
checkbox_var.set(False) 
checkbox = tk.Checkbutton(frame_mal, variable=checkbox_var,bg="#373d41", command=update_checkbox_label)
checkbox.grid(row=0, column=1, padx=0,pady=5)

checkbox_label = tk.Label(frame_mal, text="Uten bilde", fg="white", bg="#373d41")
checkbox_label.grid(row=0, column=2, padx=0)

bilde = tk.StringVar()
bilde.set("nei")

custom_style = ttk.Style()
custom_style.theme_use('alt')
custom_style.configure("Custom.TCombobox",
    font=("Roboto"),
    background="#495783",
    selectbackground="#d7d6d6",
    selectforeground="black",
    padding=(5, 1, 5, 1)
)

options = ["Piggetikett", "Hyllekant", "Stor hyllekant", "Halv A4", "Stående A4", "Liggende A4"]
selected_option = tk.StringVar()
select_box = ttk.Combobox(frame_mal, textvariable=selected_option, values=options, state="readonly",width=12, style="Custom.TCombobox")
select_box.grid(row=0, column=3, padx=5)
select_box.set(options[0])
selected_value=select_box.get()
print(selected_value)
select_box.bind("<<ComboboxSelected>>", on_combobox_select)


button3_frame = tk.Frame(root)
button3_frame.configure(background="#2b2e33")
button3_frame.pack()

run_button = tk.Button(
    button3_frame,
    text="Start",
    bg="#f15c25",
    fg="black",
    font=("Roboto",11,"bold"),
    cursor="hand2",
    command=main,
    relief=tk.FLAT,  
    borderwidth=1,
    width=11,
    height=2,  
    state=tk.DISABLED,
)
run_button.configure(cursor="hand2")
run_button.pack(pady = 10,side=tk.LEFT, padx=1)
run_button.bind("<Enter>", on_enter)
run_button.bind("<Leave>", on_leave)

save_button = tk.Button(
    button3_frame,
    text="Save",
    bg="#F1C225",
    fg="black",
    font=("Roboto",11,"bold"),
    cursor="hand2",
    command=save_data_excel,
    relief=tk.FLAT,  
    borderwidth=1,
    width=9,
    height=2,  
)
save_button.configure(cursor="hand2")
save_button.pack(pady = 10, side=tk.LEFT, padx=0)
save_button.bind("<Enter>", on_enter2)
save_button.bind("<Leave>", on_leave2)

status_label = tk.Label(root, text="")
status_label.configure(cursor="dot", fg= "#2b2e33", **label_style)
status_label.pack(pady = (0,5))


signature_label = tk.Label(root, text="Made with 🖤 by Wael for Power")
signature_label.configure(cursor="heart", bg="#2b2e33", fg="white", font=("Verdana", 10))
signature_label.pack(side="bottom", pady = 5)
signature_label.bind("<Enter>", on_enter4)
signature_label.bind("<Leave>", on_leave4)


root.mainloop()









